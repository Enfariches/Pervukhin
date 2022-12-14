import re
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
from openpyxl.styles.numbers import BUILTIN_FORMATS
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
from jinja2 import Environment, FileSystemLoader
import pdfkit
from unittest import TestCase
import unittest
import doctest

dic_naming = {'name': 'Название',
              'description': 'Описание',
              'key_skills': 'Навыки',
              'experience_id': 'Опыт работы',
              'premium': 'Премиум-вакансия',
              'employer_name': 'Компания',
              'salary_from': 'Нижняя граница вилки оклада',
              'salary_to': 'Верхняя граница вилки оклада',
              'salary_gross': 'Оклад указан до вычета налогов',
              'salary_currency': 'Идентификатор валюты оклада',
              'area_name': 'Название региона',
              'published_at': 'Дата публикации вакансии'}

currency_to_rub = {
    "AZN": 35.68,
    "BYR": 23.91,
    "EUR": 59.90,
    "GEL": 21.74,
    "KGS": 0.76,
    "KZT": 0.13,
    "RUR": 1,
    "UAH": 1.64,
    "USD": 60.66,
    "UZS": 0.0055,
}
heads1 = ["Год", "Средняя зарплата", "Средняя зарплата - ", "Количество вакансий", "Количество вакансий - "]
heads2 = ["Город", "Уровень зарплаты", '', "Город", "Доля вакансий"]

def Foo(a,b):
    return a * b

class Tests(TestCase):
    def test_clear_tag(self):
        self.assertEqual(DataSet.cleaner_string('<h>Head</h>'), 'Head')

    def test_clear_n(self):
        self.assertEqual(DataSet.cleaner_string('Head \nres \napp'), 'Head res app')

    def test_clear_many_spaces(self):
        self.assertEqual(DataSet.cleaner_string('<h>Head    res</h>'), 'Head res')

    def test_clear_spaces_sides(self):
        self.assertEqual(DataSet.cleaner_string(' <h> Head res</h> '), 'Head res')

    def test_dict_top10_less10(self):
        self.assertEqual(Report.top10({2007: 38916, 2008: 43646, 2009: 42492}),
                         {2007: 38916, 2008: 43646, 2009: 42492})

    def test_dict_top10_over10(self):
        self.assertEqual(Report.top10({2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243,
                                       2013: 51510, 2014: 50658, 2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335}),
                         {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243,
                          2013: 51510, 2014: 50658, 2015: 52696, 2016: 62675})

    def test_salary_type(self):
        self.assertEqual(type(Salary(10.0, 20.4, 'RUR')).__name__, 'Salary')

    def test_salary_currency_get_salary(self):
        self.assertEqual(Salary(10, 30.0, 'EUR').get_salary_rubles(), 1198.0)


class Report:
    """Класс создает файлы (xlsx,pdf,png) для отображения статистики вакансии, по необходимым требованиям.

    Attributes:
        report (Workbook): Переменная с функций по созданию xlxs-файла.
    """
    def __init__(self):
        """Инициализация объекта для создания xlsx-файла в корневой папке проекта.

        Args:
            report (Workbook): Переменная с функций по созданию xlxs-файла.
        """
        self.report = Report.generate_excel(result, options.parameter[1])

    @staticmethod
    def generate_excel(result, vacancy):
        """ Функция создает excel-файл с данными из статистики, оформленный по необходимым требованиям.

        Args:
            result (tuple): Статистика вакансий.
            vacancy (str): Название необходимой вакансий.

        Returns:
            Workbook: xlsx- файл, появляющийся в корневой папке проекта.
        """
        def as_text(val):
            """Функция изменяет тип объекта на str, и заменяет значение None на "" (пустоту).

            Args:
                val(str or float or int or bool or None): Переменная для перевода в str.

            Returns:
                str(val): Объект типа str.
                ""(str): Пустота типа str.
            """
            if val is None:
                return ""
            return str(val)

        def cell_parameters(sheet):
            """Функция создает рамки вокруг ячеек, и отлаживает столбец по длине загаловка.

            Args:
                sheet (sheet): xlsx лист из модуля openpyxl

            Returns:
                sheet: Измененный xlxs лист из модуля openpyxl
            """
            thin = Side(border_style="thin", color="000000")
            for column in sheet.columns:
                length = max(len(as_text(cell.value)) for cell in column)
                sheet.column_dimensions[column[0].column_letter].width = length + 2
                for cell in column:
                    if isinstance(cell.value, float):
                        cell.number_format = BUILTIN_FORMATS[10]
                    if cell.column_letter != "C" and sheet == sheet2:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                    if sheet == sheet1:
                        cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            return sheet

        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        wb = Workbook()
        sheet1 = wb.active
        sheet1.title = "Статистика по годам"
        sheet2 = wb.create_sheet("Статистика по городам")
        heads3 = [s.replace('-', f'- {vacancy}') for s in heads1]
        for i, value in enumerate(heads3):
            sheet1.cell(row=1, column=(i+1), value=value).font = Font(bold=True)
        for key, value in salary_by_years.items():
            sheet1.append([key, value, vac_salary_by_years[key], vacs_by_years[key], vac_counts_by_years[key]])
        for i, value in enumerate(heads2):
            sheet2.cell(row=1, column=(i+1), value=value).font = Font(bold=True)
        for (key, value), (k, v) in zip(salary_by_cities.items(), vacs_by_cities.items()):
            sheet2.append([key, value, '', k, v])

        cell_parameters(sheet1)
        cell_parameters(sheet2)

        return wb.save("report.xlsx")

    @staticmethod
    def graphics(result, vacancy):
        """Функция использует функций mathplotlib, чтобы сгенерировать png-формат статистики по требованиям.

        Args:
            result (tuple): Статистика вакансии, полученная из класса Interface.
            vacancy (str): Название необходимой вакансии, полученная из функций get_parameters.
        Returns:
            function: Функция, которая создает png-файл в папке проекта.
        """
    @staticmethod
    def slash(citites):
        """Функция переносит на следующую строку города в названиях которых есть тире или пробел.
        Args:
            citites (dict): Словарь со значениями статистики для преобразования.
        Returns:
            list: Список с замененными символами.
        """
        citites = [s.replace('-', '\n').replace(' ', '\n') for s in citites]
        return citites
    @staticmethod
    def top10(dict):

        """Преобразовывыет словарь для круговой диаграммы, сохраняет первые 10 пар ключ-значения,
            а остальное приводит к общему ключу ("Другие").

        Args:
            dict (dict): Словарь со значениями статистики для преобразования.
        Returns:
            dict: Преобразованный словарь.

        >>>Report.top10({2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243, 2013: 51510, 2014: 50658, 2015: 52696, 2016: 62675, 2017: 60935, 2018: 58335})
        {2007: 38916, 2008: 43646, 2009: 42492, 2010: 43846, 2011: 47451, 2012: 48243,
        2013: 51510, 2014: 50658, 2015: 52696, 2016: 62675})

        >>>Report.top10({2007: 38916, 2008: 43646, 2009: 42492})
        {2007: 38916, 2008: 43646, 2009: 42492})
        """
        first10pairs = {k: dict[k] for k in list(dict)[:11]}
        lastpairs = {k: dict[k] for k in list(dict)[10:]}
        count = 0
        for i in lastpairs.values():
            count += i
        lastpairscount = {"Другие": count}
        first10pairs.update(lastpairscount)
        return first10pairs

        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        width = 0.4
        x_nums = np.arange(len(salary_by_years.keys()))
        x_list1 = list(map(lambda x: x - width / 2, x_nums))
        x_list2 = list(map(lambda x: x + width / 2, x_nums))


        fig = plt.figure()
        ax = fig.add_subplot(221)
        ax.set_title("Уровень зарплат по годам")
        ax.bar(x_list1, salary_by_years.values(), width, label='средняя з/п')
        ax.bar(x_list2, vac_salary_by_years.values(), width, label=f'з/п {vacancy}')
        ax.set_xticks(x_nums, salary_by_years.keys(), rotation="vertical")
        ax.legend(loc='upper left', fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(222)
        ax.set_title("Количество вакансий по годам")
        ax.bar(x_list1, vacs_by_years.values(), width, label='количество вакансий')
        ax.bar(x_list2, vac_counts_by_years.values(), width, label=f'количество вакансий \n{vacancy}')
        ax.set_xticks(x_nums, vacs_by_years.keys(), rotation="vertical")
        ax.legend(loc='upper left', fontsize=8)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="y")

        ax = fig.add_subplot(223)
        ax.set_title("Уровень зарплат по городам")
        ax.barh(list(reversed(slash(salary_by_cities.keys()))), list(reversed(list(salary_by_cities.values()))))
        plt.yticks(fontsize=6,linespacing=0.66)
        ax.tick_params(axis="both", labelsize=8)
        ax.grid(True, axis="x")

        ax = fig.add_subplot(224)
        cities_finaly = top10(vacs_by_cities)
        ax.set_title("Доля вакансий по городам")
        ax.pie(list(cities_finaly.values()), labels=list(cities_finaly.keys()), textprops={'fontsize':6})
        ax.axis("equal")
        plt.tight_layout()

        return plt.savefig('graph.png', dpi=300)

    def generate_pdf(result, vacancy, heads1, heads2):
        """ Функция работает в паре с html-кодом. Чтобы преобразовать данные ввиде xlxs и png в pdf-формат.

        Args:
            result (tuple): Статистика вакансии, полученная из класса Interface.
            vacancy (str): Название необходимой вакансии, полученная из функций get_parameters.
        """
        salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities = result
        config = pdfkit.configuration(wkhtmltopdf=r'E:\apps\wkhtmltopdf\bin\wkhtmltopdf.exe')

        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        pdf_template = template.render({'vacancy': vacancy, "heads1": heads1, "salary_by_years": salary_by_years,
                                        "vac_salary_by_years": vac_salary_by_years, "vacs_by_years": vacs_by_years,
                                        "vac_counts_by_years": vac_counts_by_years, "heads2": heads2,
                                        "salary_by_cities": salary_by_cities, "vacs_by_cities": vacs_by_cities})

        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class Interface:
    """Класс принимает входящие данные от пользователя. Обрабатывает, фильтрует и возвращает (или выводит) с статистику
    по вхожденным в неё данным из csv-файла.

    Attributes:
        parameter (list): Хранит в себе данные введенные пользователем.
    """
    def __init__(self):
        """Инициализирует данные введенные пользователем.

        Args:
            parameter (list): Данные введенные пользователем(file_name, vacancy, method).
        """
        self.parameter = Interface.get_parameters()

    @staticmethod
    def get_parameters():
        """Функция получает начальные данные о файле и требованиях.

        Args:
            file_name (str): Название csv-файла.
            vacancy (str): Требуемая профессия.
            method (str): Способ вывода полученных результатов.
        """
        #file_name = input("Введите название файла: ")
        file_name = 'vacancies_medium.csv'
        #vacancy = input("Введите название профессии: ")
        vacancy = ''
        #method = input("Вакансии или Статистика: ")
        method = 'Вакансии'
        return file_name, vacancy, method

    @staticmethod
    def printing_data(dic_vacancies, vac_name, method):
        """Функция формирует статистику для её визуализаий и рассчитывает динамику необходимых требований.

        Args:
            dic_vacancies (list): Список вакансий.
            vac_name (str): Профессия введенная пользователем.
            method (str): Способ вывода полученных результатов.

        Returns:
            tuple: Кортеж со словарями, в которых хранится статистика по csv-файлу.
        """
        years = set()
        for vacancy in dic_vacancies:
            years.add(int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')))
        years = sorted(list(years))
        years = list(range(min(years), max(years) + 1))

        salary_by_years = {year: [] for year in years}
        vac_salary_by_years = {year: [] for year in years}

        vacs_by_years = {year: 0 for year in years}
        vac_counts_by_years = {year: 0 for year in years}

        for vacancy in dic_vacancies:
            year = int(datetime.strptime(vacancy.published_at, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y'))
            salary_by_years[year].append(vacancy.salary.get_salary_rubles())
            vacs_by_years[year] += 1
            if vac_name in vacancy.name:
                vac_salary_by_years[year].append(vacancy.salary.get_salary_rubles())
                vac_counts_by_years[year] += 1

        salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                           salary_by_years.items()}
        vac_salary_by_years = {key: int(sum(value) / len(value)) if len(value) != 0 else 0 for key, value in
                               vac_salary_by_years.items()}

        dic_area_name = {}
        for vacancy in dic_vacancies:
            if vacancy.area_name not in dic_area_name:
                dic_area_name[vacancy.area_name] = [vacancy.salary.get_salary_rubles()]
            else:
                dic_area_name[vacancy.area_name].append(vacancy.salary.get_salary_rubles())

        area_name_list = dic_area_name.items()
        area_name_list = [x for x in area_name_list if len(x[1]) / len(dic_vacancies) > 0.01]
        area_name_list = sorted(area_name_list, key=lambda item: sum(item[1]) / len(item[1]), reverse=True)
        salary_by_cities = {item[0]: int(sum(item[1]) / len(item[1])) for item in area_name_list[0: min(len(area_name_list), 10)]}
        vacs_dic = {}

        for vacancy in dic_vacancies:
            if vacancy.area_name in vacs_dic:
                vacs_dic[vacancy.area_name] += 1
            else:
                vacs_dic[vacancy.area_name] = 1

        vacs_counts = {x: round(y / len(dic_vacancies), 4) for x, y in vacs_dic.items()}
        vacs_counts = {k: val for k, val in vacs_counts.items() if val >= 0.01}
        vacs_by_cities = dict(sorted(vacs_counts.items(), key=lambda item: item[1], reverse=True))
        vacs_by_cities = dict(list(vacs_by_cities.items())[:10])
        if method == "Вакансии":
            print("Динамика уровня зарплат по годам:", salary_by_years)
            print("Динамика количества вакансий по годам:", vacs_by_years)
            print("Динамика уровня зарплат по годам для выбранной профессии:", vac_salary_by_years)
            print("Динамика количества вакансий по годам для выбранной профессии:", vac_counts_by_years)
            print("Уровень зарплат по городам (в порядке убывания):", salary_by_cities)
            print("Доля вакансий по городам (в порядке убывания):", vacs_by_cities)
            exit()
        elif method == "Статистика":
            dicts = salary_by_years, vac_salary_by_years, vacs_by_years, vac_counts_by_years, salary_by_cities, vacs_by_cities
            return dicts


class DataSet:
    """ Класс для получения обработанных данных csv-файла в удобном формате.

    Attributes:
        file_name (str): Введеная пользователем название csv-файла, полученная функций get_parameters.
    """
    def __init__(self, file_name):
        """ Инициализация объекта file_name, и полученного обработаного списка вакансий vacancies_objects.

        Args:
            file_name (str): Введеная пользователем название csv-файла, полученная функций get_parameters.
            vacancies_objects (list): Обработанный список вакансий.
        """
        self.file_name = file_name
        self.vacancies_objects = DataSet.csv_filter(file_name)

    @staticmethod
    def cleaner_string(text):
        """ Функция очищает строку от html-тегов и заменяет \n (знак табуляций - перенос строки) на пробел.

        Args:
             text (str): Строка из вакансий проходящая фильтрация.

        Returns:
            str: Очищенная строка.

        >>> DataSet.cleaner_string('<h>Head</h>')
        'Head'
        >>> DataSet.cleaner_string('<h>Head    res</h>')
        'Head res'
        >>> DataSet.cleaner_string(' <h> Head res</h> ')
        'Head res'
        """
        text = re.sub(r"<[^>]+>", "", text)
        text = " ".join(text.split())
        return text

    @staticmethod
    def csv_filter(file_name):
        """ Обрабатывает список с вакансиями, оставляя необходимые объекты, и чистит вакансий с пустыми ячейками.

        Args:
            file_name (str): Введеная пользователем название csv-файла, полученная функций get_parameters.

        Returns:
            list: Обработаные и отфильтрованые вакансий от html-тегов (благодаря функций cleaner_string).
        """
        list_naming, vacancies = DataSet.csv_reader(file_name)
        okay = [x for x in vacancies if len(x) == len(list_naming) and '' not in x]
        people_data = []
        dic_changed_vacancies = {}
        for line in okay:
            for i in range(len(line)):
                dic_changed_vacancies[list_naming[i]] = DataSet.cleaner_string(line[i])
            people_data.append(Vacancy(dic_changed_vacancies['name'],
                                Salary(dic_changed_vacancies['salary_from'], dic_changed_vacancies['salary_to'],
                                dic_changed_vacancies['salary_currency']),
                                dic_changed_vacancies['area_name'], dic_changed_vacancies['published_at']))
        return people_data

    @staticmethod
    def csv_reader(file_name):
        # Функция открывает и преобразовывает csv формат.

        # Args:
            # file_name (str): Введеная пользователем название csv-файла, полученная функций get_parameters.

        #Returns:
            # list (list_naming): Название столбцов csv-файла.
            # list (vacancies): Список списков всех вакансий csv-файла.

       with open(file_name, encoding="utf_8_sig") as file:
            text = csv.reader(file)
            data = [x for x in text]
            if len(data) == 0:
                print("Пустой файл")
                exit()
            list_naming = data[0]
            vacancies = data[1:]
            return list_naming, vacancies


    @staticmethod
    def test_data(arg, method):
        """Функция проверяет файл на пустоту.

        Args:
            arg(tuple): Содержит данные веденные пользователем.
            method(str): Пользовательский метод изображения результатов.

        Returns:
            tuple: Кортеж с полностью обработанными словарями.
        """
        if arg is not None:
            data = DataSet(arg[0])
            tuple_dicts = Interface.printing_data(data.vacancies_objects, arg[1], method)
            return tuple_dicts


class Salary:
    """ Класс для представления зарплаты.

    Attributes:
        salary_from (int): Нижняя граница оклада.
        salary_to (int): Верхняя граница оклада.
        salary_currency (str): Валюта оклада.
    """
    def __init__(self, salary_from, salary_to, salary_currency):
        """ Инициализирует объект Salary, выполняет конвертацию для целочисленных полей. Вычисляет среднюю зарплату
        из вилки и переводить в рубли, при помощи словоря - currency_to_rub.

        Args:
            salary_from (str or int or float): Нижняя граница оклада.
            salary_to (str or int or float): Верхняя граница оклада.
            salary_currency (str): Валюта оклада.
         >>> Salary(10000,20000,'RUR').salary_rubles
        15000.0
        >>> Salary(10000,20000,'EUR').salary_from
        599000.0
        >>> Salary(10000,20000,'EUR').salary_to
        1198000.0
        >>> Salary(10000,20000,'RUR').salary_currency
        'RUR'
        """

        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_currency = salary_currency
        self.salary_rubles = int((float(self.salary_from) + float(self.salary_to)) / 2) \
                             * currency_to_rub[self.salary_currency]

    def get_salary_rubles(self):
        """ Функция предоставляет обработанные данные о зарплате при её вызове.

        Returns:
            float: Средняя зарплата в рублях.
        """
        return self.salary_rubles

class Vacancy:
    """Класс для представления необходимых объектов вакансий.

    Attributes:
        name (str): Название вакансий.
        salary (str or int or float): Зарплата.
        area_name (str): Название региона.
        published_at (str): Дата публикаций.
    """
    def __init__(self, name, salary, area_name, published_at):
        """Инициализирует необходимые объекты name, salary, area_name, published_at и упускает ненужные объекты.

        Args:
            name (str): Название вакансии.
            salary (str or int or float): Зарплата.
            area_name (str): Название региона.
            published_at (str): Дата публикаций.
        """
        self.name = name
        self.salary = salary
        self.area_name = area_name
        self.published_at = published_at

options = Interface()
result = DataSet.test_data(options.parameter, options.parameter[2])
Report.graphics(result, options.parameter[1])
Report.generate_excel(result, options.parameter[1])
Report.generate_pdf(result, options.parameter[1], heads1, heads2)
